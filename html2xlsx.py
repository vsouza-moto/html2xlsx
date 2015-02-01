#!/usr/bin/env python

"""
    A script that extracts test suites execution results, from html files
    to xlsx spreadsheets.
"""

__author__ = "vsouza@motorola.com"
__date__ = "01/26/2015"

# importing
import argparse
import re
import os
import os.path
from os.path import getctime
import datetime
import time
import sys

import lxml.html
import openpyxl

# global variables
VERBOSE = False
RECURSIVE = False
CONSOLIDATE = False
HTML_FILENAME = None
BASENAME = None
HTML_REGEX = None

TEST_NAME_PREFIX = "TRACE LOG FOR "
TEST_FILE_REGEX = re.compile(r'(?P<test_file>[\w\-.]+\.py)')

STATUS_STRINGS = ['PASS', 'FAIL', 'ERROR']


# auxiliary functions
def clear_test_name(test_name):

    """
        Clears the test name retrieved from the html file from the trailing
        spaces and leading single quotes.
    """

    # if test_name:

    #     can_open_sq = False
    #     can_close_sq = False

    #     def accept_char(char):

    #         pass

    #     allowed_chars = [char for char in test_name.strip() if accept_char(char)]

    #     # one single quote was left unclosed, remove the 
    #     if can_close_sq:

    #         allowed_chars = allowed_chars[:-1]

    #     return "".join(allowed_chars)
    
    # else:

    #     return test_name

    return test_name.strip()


def log(message, force=False):

    """
        Logs a message depending on the VERBOSE variable current value.
        With force=True the message is logged regardless the VERBOSE value.
    """

    if VERBOSE or force:
        print ':: {0}\n'.format(message)


def scan_folder_for_html_files(folder, results):

    """
        Scans a folder for html files to extract information from.
        The scanning can be recursive or not, depending on the value
        of 'RECURSIVE'.
    """

    log("Searching for html results files in the '{0}' folder.".format(folder))

    for child in os.listdir(folder):

        pathname = os.path.join(folder, child)

        # child is a file
        if os.path.isfile(pathname):

            # check if this is a valid tests results html file
            if HTML_REGEX.match(child):

                extract_info_from_html(pathname, results)

        # child is a folder
        else:

            if RECURSIVE:

                scan_folder_for_html_files(pathname, results)


def extract_info_from_log(log_file):

    """
        Extracts the name and origin file of a test from the log file
        specific for that test.
    """

    try:

        if os.path.getsize(log_file) > 0:

            with open(log_file) as file_handler:

                test_file = None
                test_name = None

                # the first line of the log holds the test name
                for line in file_handler:

                    if test_name and test_file:

                        break

                    # if test_name is still missing
                    if not test_name:

                        prefix_index = line.find(TEST_NAME_PREFIX)

                        if prefix_index == 0:

                            test_name = line[len(TEST_NAME_PREFIX):]

                            continue

                    # if test_file is still missing
                    if not test_file:

                        match = TEST_FILE_REGEX.search(line)

                        if match:

                            test_file = match.groupdict()["test_file"]

                            continue

                # finding out what will be returned
                if test_file and test_name:

                    return (test_name, test_file)

                else:

                    if test_name:

                        return (test_name, None)

                    elif test_file:

                        return (None, test_file)

                    else:

                        return None

    except OSError:

        return None


def extract_info_from_html(html_file, results):

    """
        Extracts test round information from a single html file and then
        put the results in the 'results' map.
    """

    log("File to extract info: '{0}'".format(html_file))

    with open(html_file) as file_handler:

        html_tree = lxml.html.parse(file_handler)

        tests_names = []
        tests_status = []
        exec_times = []
        tests_files = []

        for elem in html_tree.iter():

            if elem.tag == "a":

                # if the link has text at all
                if elem.text_content():

                    # a test round status
                    if elem.text_content() in STATUS_STRINGS:

                        tests_status.append(elem.text_content())

                    else:

                        parent_dir = os.path.dirname(html_file)
                        link = os.path.join(parent_dir, elem.get("href"))

                        log("Getting the test name from file {0}".format(link))

                        test_info = extract_info_from_log(link)

                        if test_info:

                            test_real_name, test_file = test_info

                            log("Test name retrieved: {0}".format(
                                test_real_name))

                            log("Test file name retrieved: {0}".format(
                                test_file))

                            tests_names.append(clear_test_name(test_real_name))

                            tests_files.append(test_file)

                        exec_times.append(elem.tail.strip())

        if tests_names and tests_status and exec_times and tests_files:

            file_results = dict(zip(tests_names,
                                    zip(tests_status, exec_times, tests_files)
                                    ))

            results[os.path.abspath(html_file)] = file_results


def calculate_pass_rate(results):

    """
        Finds the pass rate (# of PASS tests / total # of tests) from the
        given map.
    """

    log("Calculating pass rate...")

    passed_count = reduce(
        lambda x, y: x + 1 if y == STATUS_STRINGS[0] else x,
        [info[0] for info in results.values()],
        0
    )

    return "{0:.2f}%".format(100 * (passed_count / float(len(results))))


def calculate_total_exec_time(results):

    """
        Gives the total run time of a test round.
    """

    log("Calculating total execution time...")

    total_seconds = sum([float(info[1]) for info in results.values()])

    minutes, seconds = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)

    final_time = []

    if hours != 0:

        final_time.append("{0}h".format(int(round(hours))))

    if minutes != 0:

        final_time.append("{0}m".format(int(round(minutes))))

    if seconds != 0:

        final_time.append("{0}s".format(int(round(seconds))))

    return "".join(final_time)


def consolidate_results(multiple_runs):

    """
        Consolidate the tests rounds results extracted from multiple files,
        each file holding the results from the execution of one round.
        The main objective is to have the current status of all tests - either
        'PASS', 'FAIL' or 'ERROR' - considering only the latest run of each
        test.
    """

    log("Consolidating results...")

    consolidated = dict()
    latest_test_run = dict()

    for origin, extracted in multiple_runs.items():

        creation_time = getctime(origin)

        for test_name, info in extracted.items():

            # check if the test was run before
            if test_name in latest_test_run:

                # if the current run is more recent than the one recorded
                if creation_time < latest_test_run[test_name][0]:

                    latest_test_run[test_name] = (creation_time, origin)
                    consolidated[test_name] = info
            else:

                latest_test_run[test_name] = (creation_time, origin)
                consolidated[test_name] = info

    return (consolidated, latest_test_run)


if __name__ == "__main__":

    parser = argparse.ArgumentParser()

    parser.description = __doc__

    # arguments
    parser.add_argument("input_folder",
                        nargs="*",
                        default=".",
                        help="The input folder(s) where the html files\
                        are located.")

    parser.add_argument("--output_folder",
                        "-o",
                        default=".",
                        help="The folder where the output should be put.")

    parser.add_argument("--verbose",
                        action="store_true",
                        default=False,
                        help="If the program should give more information on\
                        its progress.")

    parser.add_argument("--consolidate",
                        "-c",
                        action="store_true",
                        default=False,
                        help="Whether or not to consolidate the results from\
                        all the input folders in one spreadsheet. Conflicts\
                        regarding a same test round in different folders will\
                        be resolved by considering the result of the most\
                        recent run.")

    parser.add_argument("--recursive",
                        "-r",
                        action="store_true",
                        default=False,
                        help="Whether or not the program should search\
                        recursively for html files in the given\
                        input directories.")

    parser.add_argument("--html_filename",
                        "-f",
                        default=None,
                        help="The filename for the html files the program\
                        will try to parse (without the trailing '.html').\
                        Regular expressions are accepted.")

    parser.add_argument("--basename",
                        "-b",
                        default=None,
                        help="Usually the name of the module of feature being \
                        tested. This name will be used to compose the file \
                        name of the generated spreadsheet.")

    args = parser.parse_args()

    # retrieving the arguments
    input_folders = args.input_folder

    if not isinstance(input_folders, list):
        input_folders = [input_folders]

    OUTPUT_FOLDER = args.output_folder

    VERBOSE = args.verbose

    CONSOLIDATE = args.consolidate

    RECURSIVE = args.recursive

    if args.html_filename is None:

        HTML_REGEX = r".*\.html"

    else:

        HTML_REGEX = r"{0}\.html".format(args.html_filename)

    HTML_REGEX = re.compile(HTML_REGEX)

    BASENAME = args.basename

    # create the output folder if it doesn't exist yet
    if not os.path.exists(OUTPUT_FOLDER):

        log("Creating the '{0}' output folder...".format(OUTPUT_FOLDER))

        os.makedirs(OUTPUT_FOLDER)

    else:

        if os.path.isfile(OUTPUT_FOLDER):

            log("The '{0}' output folder can't be created! A file with the \
            same name already exists.".format(OUTPUT_FOLDER), force=True)

    tests_runs_results = dict()

    # run through all input folders and extract the data from the html files
    for input_folder in input_folders:

        if os.path.exists(input_folder):

            if os.path.isdir(input_folder):

                scan_folder_for_html_files(input_folder, tests_runs_results)

            else:

                log("The '{0}' input folder is a file!".format(input_folder),
                    force=True)

                sys.exit(1)

        else:

            log("The '{0}' input folder does not exist!".format(input_folder),
                force=True)

            sys.exit(1)

    if not tests_runs_results:

        log("No information could be extracted.", force=True)

        sys.exit(1)

    # writing the information to a .xls file

    workbook = openpyxl.Workbook(optimized_write=True)
    spreadsheet_name = None

    if BASENAME:

        spreadsheet_name = "{0} Test Round".format(BASENAME)

    else:

        spreadsheet_name = "Test Round"

    if CONSOLIDATE:

        spreadsheet_name = "{0}(consolidated)".format(spreadsheet_name)

        # consolidate the results
        extracted, latest_test_run = consolidate_results(tests_runs_results)

        # write the final data to the worksheet
        log("Creating worksheet for the consolidated results...")

        sheet = workbook.create_sheet()

        sheet.title = "Run"

        sheet.append(["Test Name",
                      "Run Status",
                      "Execution Time",
                      "Test File",
                      "Latest Run",
                      "Extracted from",
                      "",
                      "Pass Rate:",
                      calculate_pass_rate(extracted),
                      "Total Execution Time:",
                      calculate_total_exec_time(extracted)])

        extracted_items = sorted(extracted.items())

        for test_name, info in extracted_items:

            run_datetime = datetime.datetime.fromtimestamp(
                latest_test_run[test_name][0])

            str_run_datetime = run_datetime.strftime("%m/%d/%Y %H:%M:%S")

            str_run_origin = latest_test_run[test_name][1]

            sheet.append([test_name, info[0],
                         float(info[1]), info[2],
                         str_run_datetime, str_run_origin])

    else:

        round_count = 0

        sorted_results = sorted(
            tests_runs_results.items(),
            cmp=lambda x, y: int(getctime(x[0]) - getctime(y[0])))

        # for each file parsed, create a new worksheet and write the results
        for origin, extracted in sorted_results:

            log("Creating worksheet for the '{0}' results...".format(origin))

            sheet = workbook.create_sheet()

            creation_time = datetime.datetime.fromtimestamp(getctime(origin))

            time_str = creation_time.strftime("%Y%m%d %H%M")

            round_count = round_count + 1

            sheet.title = "Run #{0} {1}".format(round_count, time_str)

            sheet.append(["Test Name",
                          "Run Status",
                          "Execution Time",
                          "Test File",
                          "",
                          "Pass Rate:",
                          calculate_pass_rate(extracted),
                          "Total Execution Time:",
                          calculate_total_exec_time(extracted),
                          "Extracted from:",
                          origin])

            extracted_items = sorted(extracted.items())

            for test_name, result in extracted_items:

                sheet.append([test_name, result[0],
                             float(result[1]), result[2]])

    time_info = datetime.datetime.now().strftime("%Y%m%d %H%M")

    filename = "{0} {1}.xlsx".format(spreadsheet_name, time_info)

    filename = os.path.join(OUTPUT_FOLDER, filename)

    log("Saving the '{0}' spreadsheet at the '{1}' folder ...".format(
        spreadsheet_name,
        OUTPUT_FOLDER)
        )

    workbook.save(filename)

    log("Spreadsheet file saved")
