"""
    Writes extracted information to a XLSX spreadsheet.
"""

__author__ = "vsouza@motorola.com"
__date__ = "01/26/2015"

# importing
import openpyxl
import logging
import utils
import datetime
import os
from os.path import getctime


class Exporter(object):

    """
        This class implements the spreadsheet-writing capabilities of \
        the 'exporter' module.
    """

    def __init__(self, basename=None, output_folder="."):

        self.basename = basename
        self.output_folder = output_folder

    def consolidate(self, info):

        """
            Consolidate the tests rounds results extracted from multiple files,
            each file holding the results from the execution of one round.
            The main objective is to have the current status of all tests -
            either 'PASS', 'FAIL' or 'ERROR' - considering only the latest
            run of each test.
        """

        logging.info("Consolidating results...")

        consolidated = dict()
        latest_test_run = dict()

        for origin, extracted in info.items():

            creation_time = getctime(origin)

            for test_name, info in extracted.items():

                # check if the test was run before
                if test_name in latest_test_run:

                    # if the current run is more recent than the one recorded
                    if creation_time < latest_test_run[test_name][0]:

                        latest_test_run[test_name] = (creation_time, origin)
                        consolidated[test_name] = info
                else:

                    latest_test_run[test_name] = (creation_time, origin)
                    consolidated[test_name] = info

        return (consolidated, latest_test_run)

    def export_to_xlsx(self, info):

        """
            Exports the given data to a xlsx spreadsheet.
        """

        workbook = openpyxl.Workbook(optimized_write=True)
        spreadsheet_name = None

        if self.basename:

            spreadsheet_name = "{0} Test Round".format(self.basename)

        else:

            spreadsheet_name = "Test Round"

        round_count = 0

        comparison_function = lambda x, y: int(getctime(x[0]) - getctime(y[0]))
        sorted_results = sorted(info.items(), cmp=comparison_function)

        # for each file parsed, create a new worksheet and write the results
        for origin, extracted in sorted_results:

            logging.info("Creating worksheet for the '{0}' results...".format(origin))

            sheet = workbook.create_sheet()

            creation_time = datetime.datetime.fromtimestamp(getctime(origin))

            time_str = creation_time.strftime("%Y%m%d %H%M")

            round_count = round_count + 1

            sheet.title = "Run #{0} {1}".format(round_count, time_str)

            # appending the header line to the spreadsheet
            sheet.append(["Test Name",
                          "Test Suite",
                          "File",
                          "Function",
                          "Description",
                          "Status",
                          "Exec Time",
                          "--",
                          "Pass Rate:",
                          utils.calculate_pass_rate(extracted),
                          "Total Execution Time:",
                          utils.calculate_total_exec_time(extracted),
                          "Extracted from:",
                          origin])

            extracted_items = sorted(extracted.items())

            for test_name, info in extracted_items:

                info_list = [test_name]
                info_list.extend(info)

                sheet.append(info_list)

        time_info = datetime.datetime.now().strftime("%Y%m%d %H%M")

        filename = "{0} {1}.xlsx".format(spreadsheet_name, time_info)

        filename = os.path.join(self.output_folder, filename)

        logging.info("Saving the '{0}' spreadsheet at the '{1}' folder \
            ...".format( spreadsheet_name, self.output_folder))

        workbook.save(filename)

        logging.info("Spreadsheet file saved")

    def export_to_xlsx_consolidated(self, info):

        """
            Exports the given data to a xlsx spreadsheet, consolidating the
            results first.
        """

        workbook = openpyxl.Workbook(optimized_write=True)
        spreadsheet_name = None

        if self.basename:

            spreadsheet_name = "{0} Test Round (consolidated)".format(self.basename)

        else:

            spreadsheet_name = "Test Round (consolidated)"

        # consolidate the results
        extracted, latest_test_run = self.consolidate(info)

        # write the final data to the worksheet
        logging.info("Creating worksheet for the consolidated results...")

        sheet = workbook.create_sheet()

        sheet.title = "Run"

        sheet.append(["Test Name",
                      "Test Suite",
                      "File",
                      "Function",
                      "Description",
                      "Status",
                      "Exec Time",
                      "Latest Run",
                      "Extracted From",
                      "--",
                      "Pass Rate:",
                      utils.calculate_pass_rate(extracted),
                      "Total Execution Time:",
                      utils.calculate_total_exec_time(extracted)])

        extracted_items = sorted(extracted.items())

        for test_name, info in extracted_items:

            run_datetime = datetime.datetime.fromtimestamp(
                latest_test_run[test_name][0])

            str_run_datetime = run_datetime.strftime("%m/%d/%Y %H:%M:%S")

            str_run_origin = latest_test_run[test_name][1]

            info_list = [test_name]
            info_list.extend(info)
            info_list.extend([str_run_datetime, str_run_origin])

            sheet.append(info_list)

        time_info = datetime.datetime.now().strftime("%Y%m%d %H%M")

        filename = "{0} {1}.xlsx".format(spreadsheet_name, time_info)

        filename = os.path.join(self.output_folder, filename)

        logging.info("Saving the '{0}' spreadsheet at the '{1}' folder \
            ...".format( spreadsheet_name, self.output_folder))

        workbook.save(filename)

        logging.info("Spreadsheet file saved")
